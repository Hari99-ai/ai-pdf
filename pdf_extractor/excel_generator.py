import io
import re
from collections import Counter
from typing import Dict, Any, List, Tuple, Union
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pdf_extractor.logging_config import logger

def excel_safe_value(value: Any) -> Union[str, int, float, bool]:
    """Convert any nested JSON values into a clean string representation for Excel cells."""
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, dict):
        parts = []
        for k, v in value.items():
            val_str = str(excel_safe_value(v))
            if not val_str:
                continue
            parts.append(f"{k.replace('_', ' ').title()}: {val_str}")
        return "; ".join(parts)
    if isinstance(value, list):
        if not value:
            return ""
        # If list of dicts, format them nicely
        return ", ".join(str(excel_safe_value(item)) for item in value if item is not None)
    return str(value)

def pretty_header_name(name: str) -> str:
    """Format snake_case header name to Title Case with spaces."""
    # Special acronyms or cases
    special_cases = {
        "id": "ID",
        "usd": "USD",
        "ocr": "OCR",
        "pdf": "PDF",
        "fit": "FIT"
    }
    words = name.replace("_", " ").split()
    formatted_words = [special_cases.get(w.lower(), w.capitalize()) for w in words]
    return " ".join(formatted_words)

def union_align_sections(sections: List[Dict[str, Any]]) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Union-align multiple sections into a single list of rows with a consistent schema.
    Converts list and key_value types to tables.

    Hierarchy fields (season, date_from, date_to, meal_plan, min_stay) are always
    placed at the front of the column order so the Excel output preserves the
    Season > Date > Room > Rate hierarchy visually.
    """
    all_rows: List[Dict[str, Any]] = []
    seen_headers: List[str] = []

    # Hierarchy fields always come first, in this exact order
    HIERARCHY_FIELDS = ["season", "date_from", "date_to", "meal_plan", "min_stay"]

    for sec in sections:
        sec_type = sec.get("type", "table")
        sec_name = sec.get("name", "data")
        raw_rows = sec.get("rows", [])

        # 1. Convert key_value sections to tables (Attribute, Value)
        if sec_type == "key_value":
            if isinstance(raw_rows, dict):
                table_rows = []
                for k, v in raw_rows.items():
                    table_rows.append({
                        "attribute": pretty_header_name(k),
                        "value": excel_safe_value(v)
                    })
                raw_rows = table_rows
            elif isinstance(raw_rows, list):
                pass

        # 2. Convert list sections to tables (Description)
        elif sec_type == "list":
            if isinstance(raw_rows, list):
                raw_rows = [{"description": excel_safe_value(item)} for item in raw_rows]

        if not isinstance(raw_rows, list):
            logger.warning(f"Section {sec_name} rows is not a list. Skipping.")
            continue

        # Collect headers in order of appearance
        for row in raw_rows:
            if not isinstance(row, dict):
                continue
            for k in row.keys():
                if k not in seen_headers:
                    seen_headers.append(k)
            all_rows.append(row)

    # Sort headers: hierarchy fields first, then by priority keywords
    priority_keywords = ["date", "room", "season", "service", "description", "item", "charge", "price", "rate", "policy"]

    def header_sort_key(header: str) -> int:
        header_lower = header.lower()
        # Hierarchy fields get priority 0-4
        for idx, hf in enumerate(HIERARCHY_FIELDS):
            if header_lower == hf:
                return idx
        # Then other priority keywords
        for idx, kw in enumerate(priority_keywords):
            if kw in header_lower:
                return len(HIERARCHY_FIELDS) + idx
        return len(HIERARCHY_FIELDS) + len(priority_keywords)

    seen_headers.sort(key=header_sort_key)

    return seen_headers, all_rows


# ---------------------------------------------------------------------------
# Raw-text tabulation helpers
# ---------------------------------------------------------------------------

_PIPE_SPLIT_RE = re.compile(r"\|")
_KEY_VALUE_RE = re.compile(r"^(\S.{1,40}?)[\s]*:\s+(.+)$")
_MULTI_SPACE_RE = re.compile(r"  +")


def _is_pipe_table_row(line: str) -> bool:
    """Return True if the line looks like a pipe-delimited table row."""
    stripped = line.strip()
    if stripped.startswith("|") and stripped.endswith("|"):
        return _PIPE_SPLIT_RE.split(stripped)[1:-1] and True
    return False


def _parse_pipe_table(lines: List[str]) -> Tuple[List[str], List[Dict[str, str]]]:
    """Parse a block of pipe-delimited lines into headers and rows."""
    data_rows: List[str] = []
    separator_seen = False
    for line in lines:
        stripped = line.strip()
        # Skip separator rows like |---|---|
        if re.match(r"^\|[\s\-:|]+\|$", stripped):
            separator_seen = True
            continue
        if _is_pipe_table_row(stripped):
            data_rows.append(stripped)

    if not data_rows:
        return [], []

    def _split_pipe(line: str) -> List[str]:
        parts = line.strip().strip("|").split("|")
        return [p.strip() for p in parts]

    # Use first row as header
    headers = _split_pipe(data_rows[0])
    rows: List[Dict[str, str]] = []
    for line in data_rows[1:]:
        cells = _split_pipe(line)
        row = {}
        for i, h in enumerate(headers):
            row[h if h else f"col_{i+1}"] = cells[i] if i < len(cells) else ""
        rows.append(row)

    return headers, rows


def _detect_aligned_columns(lines: List[str]) -> Tuple[List[int], List[str]]:
    """Detect column boundaries by finding consistent multi-space gaps across lines.

    Returns (cut_positions, headers_derived_from_positions).
    """
    # Find all gap positions (2+ spaces) in each line
    gap_positions_by_line: List[List[int]] = []
    for line in lines:
        gaps = [m.start() for m in _MULTI_SPACE_RE.finditer(line)]
        gap_positions_by_line.append(gaps)

    if len(gap_positions_by_line) < 2:
        return [], []

    # Count how often each position appears across lines
    position_counter: Counter = Counter()
    for gaps in gap_positions_by_line:
        for pos in gaps:
            position_counter[pos] += 1

    # Keep positions that appear in at least 40% of lines
    threshold = max(2, len(lines) * 0.4)
    candidate_positions = sorted(pos for pos, count in position_counter.items() if count >= threshold)

    if len(candidate_positions) < 1:
        return [], []

    # Merge positions that are very close (within 3 chars) – keep the earliest
    merged: List[int] = []
    for pos in candidate_positions:
        if merged and pos - merged[-1] < 3:
            continue
        merged.append(pos)

    return merged, []


def _split_by_positions(line: str, positions: List[int]) -> List[str]:
    """Split a line at fixed column positions."""
    cells: List[str] = []
    start = 0
    for pos in positions:
        if pos > len(line):
            cells.append(line[start:].strip())
            start = len(line)
            break
        cells.append(line[start:pos].strip())
        start = pos
    cells.append(line[start:].strip())
    return [c for c in cells if c]


def _tabulate_aligned_columns(lines: List[str]) -> Tuple[List[str], List[Dict[str, str]]]:
    """Attempt to detect space-aligned columns and parse into headers + rows."""
    positions, _ = _detect_aligned_columns(lines)
    if not positions:
        return [], []

    # Split all lines
    split_lines = [_split_by_positions(line, positions) for line in lines]

    # Heuristic: if first line has more cells than most others, treat it as header
    cell_counts = [len(cells) for cells in split_lines]
    if not cell_counts:
        return [], []

    most_common_count = Counter(cell_counts).most_common(1)[0][0]
    header_line = split_lines[0]
    data_lines = split_lines[1:]

    # If the first line has a different number of columns, it's likely a header
    if len(header_line) != most_common_count and len(data_lines) > 0:
        header_cells = header_line
        data_split = data_lines
    else:
        header_cells = [f"Column {i+1}" for i in range(most_common_count)]
        data_split = split_lines

    # Pad header if needed
    while len(header_cells) < most_common_count:
        header_cells.append(f"Column {len(header_cells)+1}")

    headers = header_cells[:most_common_count]
    rows: List[Dict[str, str]] = []
    for cells in data_split:
        row = {}
        for i, h in enumerate(headers):
            row[h] = cells[i] if i < len(cells) else ""
        # Skip fully empty rows
        if any(v.strip() for v in row.values()):
            rows.append(row)

    return headers, rows


def _tabulate_key_value(lines: List[str]) -> Tuple[List[str], List[Dict[str, str]]]:
    """Parse lines that look like 'Label: Value' into an Attribute / Value table."""
    rows: List[Dict[str, str]] = []
    for line in lines:
        m = _KEY_VALUE_RE.match(line.strip())
        if m:
            rows.append({"attribute": m.group(1).strip(), "value": m.group(2).strip()})

    if not rows:
        return [], []
    return ["attribute", "value"], rows


def tabulate_raw_text(raw_text: str) -> Tuple[List[str], List[Dict[str, str]]]:
    """Parse raw PDF text into structured (headers, rows) suitable for an Excel sheet.

    Detection order:
    1. Pipe-delimited tables (| col | col |)
    2. Space-aligned columns
    3. Key-value pairs (Label: Value)
    4. Plain text lines as single-column fallback
    """
    if not raw_text or not raw_text.strip():
        return [], []

    lines = [line.rstrip() for line in raw_text.splitlines()]
    non_empty = [line for line in lines if line.strip()]

    if not non_empty:
        return [], []

    # --- Strategy 1: Pipe-delimited table ---
    pipe_lines = [line for line in non_empty if _is_pipe_table_row(line)]
    if len(pipe_lines) >= 2:
        logger.info(f"Detected pipe-delimited table with {len(pipe_lines)} rows.")
        return _parse_pipe_table(pipe_lines)

    # --- Strategy 2: Space-aligned columns ---
    # Use only lines that have at least 2 multi-space gaps (likely columnar data)
    colar_lines = [line for line in non_empty if len(_MULTI_SPACE_RE.findall(line)) >= 2]
    if len(colar_lines) >= 3:
        headers, rows = _tabulate_aligned_columns(colar_lines)
        if headers and rows:
            logger.info(f"Detected space-aligned table with {len(rows)} rows and {len(headers)} columns.")
            return headers, rows

    # --- Strategy 3: Key-value pairs ---
    kv_lines = [line for line in non_empty if _KEY_VALUE_RE.match(line.strip())]
    if len(kv_lines) >= 2:
        headers, rows = _tabulate_key_value(kv_lines)
        if rows:
            logger.info(f"Detected key-value table with {len(rows)} rows.")
            return headers, rows

    # --- Strategy 4: Plain text fallback – single "Text" column ---
    logger.info("No structured pattern detected; using single-column fallback.")
    return ["text"], [{"text": line} for line in non_empty]


def build_excel_workbook(groups: Dict[str, Dict[str, Any]], raw_pdf_text: str = "") -> bytes:
    """Build the single Excel workbook with styled worksheets and custom styles.

    Generates 3 classified group sheets plus an optional 4th sheet containing the
    raw PDF text parsed into tabulated format.
    """
    logger.info("Initializing openpyxl Workbook...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # We must process exactly group_1, group_2, group_3
    keys = ["group_1", "group_2", "group_3"]
    
    for key in keys:
        group_data = groups.get(key, {})
        sheet_title = group_data.get("sheet_title", pretty_header_name(key))
        sections = group_data.get("sections", [])
        
        ws = wb.create_sheet(title=sheet_title)
        
        # Ensure grid lines are visible
        ws.views.sheetView[0].showGridLines = True
        
        # Align sections into a single relational table
        headers, rows = union_align_sections(sections)
        
        # If no data found, insert placeholder
        if not rows:
            headers = ["status"]
            rows = [{"status": "No data detected for this category."}]
            logger.info(f"No data for sheet '{sheet_title}', creating placeholder row.")
            
        # Write headers
        ws.append([pretty_header_name(h) for h in headers])
        
        # Write rows
        for row in rows:
            ws.append([excel_safe_value(row.get(h, "")) for h in headers])
            
        # Freeze top row
        ws.freeze_panes = "A2"
        
        # Style cells
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        border_thin = Side(border_style="thin", color="D3D3D3")
        cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        
        # Center alignment for header
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Left alignment with wrap text for data cells
        align_data_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Center alignment for specific columns (dates, rates, statuses, codes)
        align_data_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Apply header formatting
        ws.row_dimensions[1].height = 26
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_header
            cell.border = cell_border
            
        # Apply data formatting
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
            # Zebra striping (alternating white and light blue)
            row_fill = PatternFill(
                start_color="F2F6FA" if row_idx % 2 == 0 else "FFFFFF",
                end_color="F2F6FA" if row_idx % 2 == 0 else "FFFFFF",
                fill_type="solid"
            )

            # Hierarchy columns get a distinct light teal background
            hierarchy_fill = PatternFill(
                start_color="E8F5E9" if row_idx % 2 == 0 else "F1F8E9",
                end_color="E8F5E9" if row_idx % 2 == 0 else "F1F8E9",
                fill_type="solid"
            )
            HIERARCHY_COLS = {"season", "date_from", "date_to", "meal_plan", "min_stay"}

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = cell_border

                h_lower = header.lower()

                # Hierarchy columns get teal fill + bold font
                if h_lower in HIERARCHY_COLS:
                    cell.fill = hierarchy_fill
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="2E7D32")
                    cell.alignment = align_data_center
                elif h_lower == "context":
                    # Style the Context column distinctly
                    cell.font = Font(name="Segoe UI", size=10, italic=True, color="555555")
                    cell.fill = PatternFill(
                        start_color="FFF8E1" if row_idx % 2 == 0 else "FFFDF5",
                        end_color="FFF8E1" if row_idx % 2 == 0 else "FFFDF5",
                        fill_type="solid"
                    )
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.fill = row_fill
                    if any(kw in h_lower for kw in ["date", "rate", "price", "status", "id", "percent"]):
                        cell.alignment = align_data_center
                    else:
                        cell.alignment = align_data_left
                    
        # Apply Excel Table format
        try:
            # Table name must be alphanumeric and start with a letter, no spaces
            clean_name = re.sub(r"[^a-zA-Z0-9]", "", sheet_title)
            if not clean_name or not clean_name[0].isalpha():
                clean_name = f"Table{clean_name}"
            # Ensure name is unique inside workbook (append sheet index if needed)
            clean_name = clean_name[:250]
            
            table_ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
            excel_table = Table(displayName=clean_name, ref=table_ref)
            
            # Medium style with row stripes
            table_style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            excel_table.tableStyleInfo = table_style
            ws.add_table(excel_table)
            logger.info(f"Added Excel Table '{clean_name}' for range {table_ref}.")
        except Exception as e:
            logger.warning(f"Could not add native Excel table styling: {e}. Falling back to cell styles.")
            
        # Enable Auto Filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        
        # Auto column width
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or "")
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    logger.info("Workbook generation finished.")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
